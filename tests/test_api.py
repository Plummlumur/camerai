from fastapi.testclient import TestClient

from api.main import create_app
from config import Settings


def make_client(tmp_path) -> TestClient:
    settings = Settings(_env_file=None, counter_source="none", db_path=str(tmp_path / "test.db"))
    return TestClient(create_app(settings))


def test_status_starts_empty(tmp_path):
    with make_client(tmp_path) as client:
        response = client.get("/api/status")
    assert response.status_code == 200
    body = response.json()
    assert body["occupancy"] == 0
    assert body["today_in"] == 0
    assert body["today_out"] == 0
    assert body["sensor_id"] == "raum-1"


def test_status_reports_camera_preview_disabled_by_default(tmp_path):
    with make_client(tmp_path) as client:
        body = client.get("/api/status").json()
    assert body["preview_enabled"] is False
    assert body["line_position"] == 0.5
    assert body["line_axis"] == "x"


def test_camera_stream_returns_404_when_preview_disabled(tmp_path):
    with make_client(tmp_path) as client:
        response = client.get("/api/camera/stream")
    assert response.status_code == 404


def test_preview_flag_ignored_for_non_camera_source(tmp_path):
    # Enabling the preview without the imx500 source must not advertise a
    # preview that can never stream an image.
    settings = Settings(
        _env_file=None,
        counter_source="none",
        camera_preview_enabled=True,
        db_path=str(tmp_path / "test.db"),
    )
    with TestClient(create_app(settings)) as client:
        assert client.get("/api/status").json()["preview_enabled"] is False
        assert client.get("/api/camera/stream").status_code == 404


def test_status_restores_occupancy_from_events(tmp_path):
    settings = Settings(_env_file=None, counter_source="none", db_path=str(tmp_path / "test.db"))
    from storage.events import EventStore
    from timeutils import utc_now_iso

    store = EventStore(settings.db_path)
    store.add_event(utc_now_iso(), "in", "raum-1")
    store.add_event(utc_now_iso(), "in", "raum-1")
    store.close()
    with TestClient(create_app(settings)) as client:
        body = client.get("/api/status").json()
    assert body["occupancy"] == 2
    assert body["today_in"] == 2


def test_dashboard_is_served(tmp_path):
    with make_client(tmp_path) as client:
        response = client.get("/")
    assert response.status_code == 200
    assert "Raumzähler" in response.text


def test_today_stats_has_24_hour_buckets(tmp_path):
    with make_client(tmp_path) as client:
        body = client.get("/api/stats/today").json()
    assert len(body["hours"]) == 24
    assert body["hours"][0] == {"hour": 0, "in": 0, "out": 0}


def test_history_defaults_to_seven_days(tmp_path):
    with make_client(tmp_path) as client:
        body = client.get("/api/stats/history").json()
    assert len(body["days"]) == 7


def test_correction_sets_occupancy_and_writes_event(tmp_path):
    with make_client(tmp_path) as client:
        response = client.post("/api/occupancy", json={"value": 7})
        assert response.status_code == 200
        assert response.json()["occupancy"] == 7
        assert client.get("/api/status").json()["occupancy"] == 7
        store = client.app.state.store
        rows = store._conn.execute("SELECT direction, value FROM events ORDER BY id").fetchall()
    assert rows == [("correction", 7)]


def test_correction_rejects_negative_values(tmp_path):
    with make_client(tmp_path) as client:
        response = client.post("/api/occupancy", json={"value": -1})
    assert response.status_code == 422


def test_history_rejects_out_of_range_days(tmp_path):
    with make_client(tmp_path) as client:
        assert client.get("/api/stats/history?days=0").status_code == 422
        assert client.get("/api/stats/history?days=367").status_code == 422


def test_history_period_returns_range_bounds(tmp_path):
    with make_client(tmp_path) as client:
        body = client.get("/api/stats/history?period=last_week").json()
    assert body["period"] == "last_week"
    # Monday..Sunday, inclusive, one daily bucket each.
    assert len(body["days"]) == 7
    assert body["days"][0]["date"] == body["start"]
    assert body["days"][-1]["date"] == body["end"]


def test_history_rejects_unknown_period(tmp_path):
    with make_client(tmp_path) as client:
        assert client.get("/api/stats/history?period=nonsense").status_code == 422


def test_export_xlsx_returns_workbook_with_five_named_sheets(tmp_path):
    import io

    from openpyxl import load_workbook

    with make_client(tmp_path) as client:
        response = client.get("/api/export/xlsx")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    assert response.headers["content-disposition"].startswith("attachment; filename=")
    wb = load_workbook(io.BytesIO(response.content))
    assert wb.sheetnames == [
        "Gestern",
        "Laufende Woche",
        "Letzte Woche",
        "Laufender Monat",
        "Letzter Monat",
    ]


def test_correction_is_broadcast_to_websocket_clients(tmp_path):
    with make_client(tmp_path) as client:
        with client.websocket_connect("/ws") as websocket:
            client.post("/api/occupancy", json={"value": 4})
            message = websocket.receive_json()
    assert message["type"] == "correction"
    assert message["occupancy"] == 4
