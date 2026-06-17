import io
from datetime import date
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from api.export import build_history_workbook
from storage.events import EventStore

VIENNA = ZoneInfo("Europe/Vienna")
EXPECTED_SHEETS = ["Gestern", "Laufende Woche", "Letzte Woche", "Laufender Monat", "Letzter Monat"]


def _load(store, today):
    return load_workbook(io.BytesIO(build_history_workbook(store, VIENNA, today)))


def test_workbook_has_five_period_sheets_in_order(tmp_path):
    store = EventStore(tmp_path / "t.db")
    wb = _load(store, date(2026, 6, 17))
    assert wb.sheetnames == EXPECTED_SHEETS
    store.close()


def test_yesterday_sheet_has_header_data_and_total(tmp_path):
    store = EventStore(tmp_path / "t.db")
    # 2026-06-16 is "Gestern" relative to 2026-06-17; events inside the local day.
    store.add_event("2026-06-16T10:00:00+00:00", "in", "s")
    store.add_event("2026-06-16T11:00:00+00:00", "in", "s")
    store.add_event("2026-06-16T12:00:00+00:00", "out", "s")
    ws = _load(store, date(2026, 6, 17))["Gestern"]
    assert [c.value for c in ws[1]] == ["Datum", "Eintritte", "Austritte"]
    assert ws.cell(2, 1).value.date() == date(2026, 6, 16)  # single day row
    assert [ws.cell(2, 2).value, ws.cell(2, 3).value] == [2, 1]
    # Totals are live SUM formulas, not precomputed values (one day -> B2:B2).
    assert [c.value for c in ws[ws.max_row]] == ["Summe", "=SUM(B2:B2)", "=SUM(C2:C2)"]
    store.close()


def test_last_week_sheet_spans_seven_days(tmp_path):
    store = EventStore(tmp_path / "t.db")
    ws = _load(store, date(2026, 6, 17))["Letzte Woche"]
    # header + 7 day rows (Mon..Sun) + Summe; the SUM spans exactly the 7 rows.
    assert ws.max_row == 9
    assert [c.value for c in ws[ws.max_row]] == ["Summe", "=SUM(B2:B8)", "=SUM(C2:C8)"]
    store.close()
