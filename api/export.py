import io
from datetime import date
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font

from storage.events import EventStore
from timeutils import period_bounds

# Worksheet order and German titles for the history export. Same periods as the
# dashboard buttons, one sheet each.
EXPORT_PERIODS = [
    ("yesterday", "Gestern"),
    ("current_week", "Laufende Woche"),
    ("last_week", "Letzte Woche"),
    ("current_month", "Laufender Monat"),
    ("last_month", "Letzter Monat"),
]


def build_history_workbook(store: EventStore, tz: ZoneInfo, today: date) -> bytes:
    """Build an .xlsx workbook with one sheet of daily in/out totals per period."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the auto-created default sheet
    bold = Font(bold=True)
    for key, title in EXPORT_PERIODS:
        start, end = period_bounds(key, today)
        ws = wb.create_sheet(title=title)
        ws.append(["Datum", "Eintritte", "Austritte"])
        rows = store.daily_totals_range(start, end, tz)
        for row in rows:
            ws.append([date.fromisoformat(row["date"]), row["in"], row["out"]])
        # Totals as live SUM formulas (not precomputed) so manual edits to the
        # daily rows recalculate. Data spans rows 2..(1+len); there is always at
        # least one day, so the range is never empty.
        first, last = 2, 1 + len(rows)
        ws.append(["Summe", f"=SUM(B{first}:B{last})", f"=SUM(C{first}:C{last})"])
        for cell in ws[1] + ws[ws.max_row]:  # bold header and total rows
            cell.font = bold
        ws.column_dimensions["A"].width = 14
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
